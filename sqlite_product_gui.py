import sys
import sqlite3
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QMessageBox, QFileDialog
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QColor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


class ProductManager:
    """SQLite를 사용하여 제품 데이터를 관리하는 클래스"""
    
    def __init__(self, db_name: str = "products.db"):
        self.db_name = db_name
        self.conn = None
        self.cursor = None
        self.connect()
        self.create_table()
    
    def connect(self):
        """데이터베이스에 연결"""
        try:
            self.conn = sqlite3.connect(self.db_name)
            self.cursor = self.conn.cursor()
        except sqlite3.Error as e:
            print(f"데이터베이스 연결 오류: {e}")
    
    def create_table(self):
        """Products 테이블 생성"""
        try:
            sql = """
            CREATE TABLE IF NOT EXISTS Products (
                productID INTEGER PRIMARY KEY AUTOINCREMENT,
                productName TEXT NOT NULL,
                productPrice INTEGER NOT NULL
            )
            """
            self.cursor.execute(sql)
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"테이블 생성 오류: {e}")
    
    def insert_product(self, product_name: str, product_price: int) -> bool:
        """제품 데이터 삽입"""
        try:
            sql = "INSERT INTO Products (productName, productPrice) VALUES (?, ?)"
            self.cursor.execute(sql, (product_name, product_price))
            self.conn.commit()
            return True
        except sqlite3.Error as e:
            print(f"제품 삽입 오류: {e}")
            return False
    
    def select_all_products(self):
        """모든 제품 조회"""
        try:
            sql = "SELECT * FROM Products ORDER BY productID"
            self.cursor.execute(sql)
            products = self.cursor.fetchall()
            return products
        except sqlite3.Error as e:
            print(f"조회 오류: {e}")
            return []
    
    def update_product(self, product_id: int, product_name: str, product_price: int) -> bool:
        """제품 정보 수정"""
        try:
            sql = "UPDATE Products SET productName = ?, productPrice = ? WHERE productID = ?"
            self.cursor.execute(sql, (product_name, product_price, product_id))
            self.conn.commit()
            return self.cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"수정 오류: {e}")
            return False
    
    def delete_product(self, product_id: int) -> bool:
        """제품 삭제"""
        try:
            sql = "DELETE FROM Products WHERE productID = ?"
            self.cursor.execute(sql, (product_id,))
            self.conn.commit()
            return self.cursor.rowcount > 0
        except sqlite3.Error as e:
            print(f"삭제 오류: {e}")
            return False
    
    def close(self):
        """데이터베이스 연결 종료"""
        if self.conn:
            self.conn.close()


class ProductGUI(QMainWindow):
    """제품 관리 GUI 애플리케이션"""
    
    def __init__(self):
        super().__init__()
        self.manager = ProductManager()
        self.selected_row = -1
        self.init_ui()
        self.apply_styles()
    
    def init_ui(self):
        """UI 초기화"""
        self.setWindowTitle("제품 관리 시스템")
        self.setGeometry(100, 100, 800, 600)
        
        # 중앙 위젯
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 메인 레이아웃
        main_layout = QVBoxLayout()
        
        # ========== 입력 영역 ==========
        input_layout = QHBoxLayout()
        
        # 제품명 입력
        label_name = QLabel("제품명:")
        self.input_name = QLineEdit()
        self.input_name.setPlaceholderText("제품명을 입력하세요")
        input_layout.addWidget(label_name)
        input_layout.addWidget(self.input_name)
        
        # 제품 가격 입력
        label_price = QLabel("가격:")
        self.input_price = QLineEdit()
        self.input_price.setPlaceholderText("가격을 입력하세요")
        input_layout.addWidget(label_price)
        input_layout.addWidget(self.input_price)
        
        main_layout.addLayout(input_layout)
        
        # ========== 버튼 영역 ==========
        button_layout = QHBoxLayout()
        
        # 추가 버튼
        btn_add = QPushButton("추가")
        btn_add.clicked.connect(self.add_product)
        button_layout.addWidget(btn_add)
        
        # 수정 버튼
        btn_update = QPushButton("수정")
        btn_update.clicked.connect(self.update_product)
        button_layout.addWidget(btn_update)
        
        # 삭제 버튼
        btn_delete = QPushButton("삭제")
        btn_delete.clicked.connect(self.delete_product)
        button_layout.addWidget(btn_delete)
        
        # 새로고침 버튼
        btn_refresh = QPushButton("새로고침")
        btn_refresh.clicked.connect(self.refresh_table)
        button_layout.addWidget(btn_refresh)
        
        # 엑셀 저장 버튼
        btn_export = QPushButton("엑셀 저장")
        btn_export.clicked.connect(self.export_to_excel)
        button_layout.addWidget(btn_export)
        
        # 초기화 버튼
        btn_clear = QPushButton("입력 초기화")
        btn_clear.clicked.connect(self.clear_input)
        button_layout.addWidget(btn_clear)
        
        main_layout.addLayout(button_layout)
        
        # ========== 테이블 영역 ==========
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["ID", "제품명", "가격"])
        self.table.setColumnWidth(0, 80)
        self.table.setColumnWidth(1, 300)
        self.table.setColumnWidth(2, 150)
        self.table.itemDoubleClicked.connect(self.on_table_double_clicked)
        
        main_layout.addWidget(self.table)
        
        # 레이아웃 설정
        central_widget.setLayout(main_layout)
        
        # 초기 데이터 로드
        self.refresh_table()
    
    def add_product(self):
        """제품 추가"""
        product_name = self.input_name.text().strip()
        product_price_str = self.input_price.text().strip()
        
        # 유효성 검사
        if not product_name:
            QMessageBox.warning(self, "경고", "제품명을 입력하세요.")
            return
        
        if not product_price_str:
            QMessageBox.warning(self, "경고", "가격을 입력하세요.")
            return
        
        try:
            product_price = int(product_price_str)
            if product_price < 0:
                QMessageBox.warning(self, "경고", "가격은 0 이상이어야 합니다.")
                return
        except ValueError:
            QMessageBox.warning(self, "경고", "가격은 숫자로 입력하세요.")
            return
        
        # 데이터베이스에 추가
        if self.manager.insert_product(product_name, product_price):
            QMessageBox.information(self, "성공", f"제품 '{product_name}'이 추가되었습니다.")
            self.clear_input()
            self.refresh_table()
        else:
            QMessageBox.critical(self, "오류", "제품 추가에 실패했습니다.")
    
    def update_product(self):
        """제품 수정"""
        if self.selected_row < 0:
            QMessageBox.warning(self, "경고", "수정할 제품을 선택하세요.")
            return
        
        product_name = self.input_name.text().strip()
        product_price_str = self.input_price.text().strip()
        
        # 유효성 검사
        if not product_name:
            QMessageBox.warning(self, "경고", "제품명을 입력하세요.")
            return
        
        if not product_price_str:
            QMessageBox.warning(self, "경고", "가격을 입력하세요.")
            return
        
        try:
            product_price = int(product_price_str)
            if product_price < 0:
                QMessageBox.warning(self, "경고", "가격은 0 이상이어야 합니다.")
                return
        except ValueError:
            QMessageBox.warning(self, "경고", "가격은 숫자로 입력하세요.")
            return
        
        # 선택된 행에서 ID 가져오기
        product_id = int(self.table.item(self.selected_row, 0).text())
        
        # 데이터베이스에서 수정
        if self.manager.update_product(product_id, product_name, product_price):
            QMessageBox.information(self, "성공", f"제품 ID {product_id}이 수정되었습니다.")
            self.clear_input()
            self.refresh_table()
        else:
            QMessageBox.critical(self, "오류", "제품 수정에 실패했습니다.")
    
    def delete_product(self):
        """제품 삭제"""
        if self.selected_row < 0:
            QMessageBox.warning(self, "경고", "삭제할 제품을 선택하세요.")
            return
        
        # 삭제 확인
        reply = QMessageBox.question(
            self, 
            "삭제 확인", 
            "선택한 제품을 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
        
        # 선택된 행에서 ID 가져오기
        product_id = int(self.table.item(self.selected_row, 0).text())
        
        # 데이터베이스에서 삭제
        if self.manager.delete_product(product_id):
            QMessageBox.information(self, "성공", f"제품 ID {product_id}이 삭제되었습니다.")
            self.clear_input()
            self.refresh_table()
        else:
            QMessageBox.critical(self, "오류", "제품 삭제에 실패했습니다.")
    
    def refresh_table(self):
        """테이블 새로고침"""
        products = self.manager.select_all_products()
        self.table.setRowCount(len(products))
        
        for row, (product_id, product_name, product_price) in enumerate(products):
            # ID 셀
            item_id = QTableWidgetItem(str(product_id))
            item_id.setAlignment(Qt.AlignCenter)
            self.table.setItem(row, 0, item_id)
            
            # 제품명 셀
            item_name = QTableWidgetItem(product_name)
            item_name.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.table.setItem(row, 1, item_name)
            
            # 가격 셀
            item_price = QTableWidgetItem(f"{product_price:,}원")
            item_price.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(row, 2, item_price)
    
    def on_table_double_clicked(self, item):
        """테이블 더블클릭 시 처리"""
        self.selected_row = self.table.row(item)
        
        # 선택된 행의 데이터를 입력 필드에 채우기
        product_name = self.table.item(self.selected_row, 1).text()
        product_price = self.table.item(self.selected_row, 2).text().replace("원", "").replace(",", "")
        
        self.input_name.setText(product_name)
        self.input_price.setText(product_price)
    
    def clear_input(self):
        """입력 필드 초기화"""
        self.input_name.clear()
        self.input_price.clear()
        self.selected_row = -1
        self.table.clearSelection()
    
    def export_to_excel(self):
        """데이터를 엑셀 파일로 저장"""
        products = self.manager.select_all_products()
        
        if not products:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다.")
            return
        
        # 파일 저장 경로 설정
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "엑셀 파일 저장",
            f"제품목록_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            # Workbook 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "제품 목록"
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 헤더 작성
            headers = ["ID", "제품명", "가격"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 데이터 작성
            for row_num, (product_id, product_name, product_price) in enumerate(products, 2):
                ws.cell(row=row_num, column=1, value=product_id)
                ws.cell(row=row_num, column=2, value=product_name)
                ws.cell(row=row_num, column=3, value=product_price)
                
                # 가격 셀을 오른쪽 정렬
                ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="right")
            
            # 열 너비 자동 조정
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            
            # 파일 저장
            wb.save(file_path)
            QMessageBox.information(self, "성공", f"엑셀 파일이 저장되었습니다.\n경로: {file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 파일 저장에 실패했습니다.\n{str(e)}")
    
    def apply_styles(self):
        """화려한 스타일 적용"""
        # 전체 윈도우 스타일
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            
            QLabel {
                font-size: 11px;
                font-weight: bold;
                color: #333333;
            }
            
            QLineEdit {
                border: 2px solid #4472C4;
                border-radius: 5px;
                padding: 8px;
                font-size: 11px;
                background-color: white;
                selection-background-color: #4472C4;
            }
            
            QLineEdit:focus {
                border: 2px solid #2E5C8A;
                background-color: #F0F4FF;
            }
            
            QPushButton {
                background-color: #4472C4;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 11px;
                font-weight: bold;
                min-height: 30px;
            }
            
            QPushButton:hover {
                background-color: #2E5C8A;
            }
            
            QPushButton:pressed {
                background-color: #1A3A5C;
            }
            
            QTableWidget {
                border: 2px solid #4472C4;
                border-radius: 5px;
                gridline-color: #D3D3D3;
                background-color: white;
            }
            
            QTableWidget::item {
                padding: 5px;
                border: none;
            }
            
            QTableWidget::item:selected {
                background-color: #E8F0FF;
            }
            
            QHeaderView::section {
                background-color: #4472C4;
                color: white;
                padding: 6px;
                border: none;
                font-weight: bold;
                font-size: 11px;
            }
            
            QMessageBox {
                background-color: #f5f5f5;
            }
            
            QMessageBox QLabel {
                color: #333333;
            }
        """)
        
        # 제목 폰트 설정
        title_font = QFont()
        title_font.setPointSize(10)
        title_font.setBold(True)
        self.setFont(title_font)
    
    def closeEvent(self, event):
        """윈도우 종료 시 처리"""
        self.manager.close()
        event.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ProductGUI()
    window.show()
    sys.exit(app.exec_())
